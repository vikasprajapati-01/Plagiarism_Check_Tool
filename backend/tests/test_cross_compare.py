"""Tests for the cross-comparison service."""

import io
import pytest
from openpyxl import Workbook

from app.services.cross_compare import (
    parse_excel_file,
    compare_rows,
    compare_cells,
    run_cross_comparison,
    generate_comparison_report,
    generate_colored_workbook,
)


def _make_test_workbook() -> bytes:
    """Create a test workbook with 2 sheets and known duplicates."""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Dataset 1"
    ws1.append(["S. No.", "Query", "Location", "Time"])
    ws1.append([1, "John playing foosball in Bangalore yesterday", "Bangalore", "yesterday"])
    ws1.append([2, "The concert I watched last week in Delhi", "Delhi", "last week"])
    ws1.append([3, "Photos of cats on the streets of Ooty", "Ooty", "this month"])
    ws1.append([4, "John playing foosball in Bangalore yesterday", "Bangalore", "yesterday"])  # exact dup of row 2
    ws1.append([5, "Japan 2025 trip photos", "Japan", "2025"])

    ws2 = wb.create_sheet("Dataset 2")
    ws2.append(["S. No.", "Query", "Location", "Time"])
    ws2.append([1, "Joel playing table tennis in Mangalore yesterday", "Mangalore", "yesterday"])
    ws2.append([2, "The concert I watched last week in Chennai", "Chennai", "last week"])
    ws2.append([3, "Tokyo cherry blossom photos from March", "Tokyo", "March"])
    ws2.append([4, "Japan 2025 trip photos", "Japan", "2025"])  # exact dup across sheets

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseExcelFile:
    def test_parses_all_sheets(self):
        contents = _make_test_workbook()
        rows, cells = parse_excel_file("test.xlsx", contents)
        # 5 rows from sheet1 + 4 rows from sheet2
        assert len(rows) == 9

    def test_skips_serial_number_column(self):
        contents = _make_test_workbook()
        rows, cells = parse_excel_file("test.xlsx", contents)
        # No cell should have a pure numeric serial number as its value
        for cell in cells:
            assert cell.col_letter != "A"  # S. No. column skipped

    def test_cells_have_references(self):
        contents = _make_test_workbook()
        rows, cells = parse_excel_file("test.xlsx", contents)
        for cell in cells:
            assert cell.file_name == "test.xlsx"
            assert cell.sheet_name in ("Dataset 1", "Dataset 2")
            assert cell.row >= 2  # header skipped


class TestCompareRows:
    def test_finds_exact_row_duplicates(self):
        contents = _make_test_workbook()
        rows, _ = parse_excel_file("test.xlsx", contents)
        matches = compare_rows(rows, threshold=75.0)

        exact_matches = [m for m in matches if m.match_type == "Exact"]
        assert len(exact_matches) >= 1  # At least row 2 ↔ row 5 in Sheet1

    def test_finds_near_row_duplicates(self):
        contents = _make_test_workbook()
        rows, _ = parse_excel_file("test.xlsx", contents)
        matches = compare_rows(rows, threshold=75.0)

        near_matches = [m for m in matches if m.match_type == "Near"]
        assert len(near_matches) >= 1  # Delhi vs Chennai concerts

    def test_threshold_filtering(self):
        contents = _make_test_workbook()
        rows, _ = parse_excel_file("test.xlsx", contents)

        strict = compare_rows(rows, threshold=95.0)
        loose = compare_rows(rows, threshold=50.0)
        assert len(loose) >= len(strict)


class TestCompareCells:
    def test_finds_exact_cell_matches(self):
        contents = _make_test_workbook()
        _, cells = parse_excel_file("test.xlsx", contents)
        matches = compare_cells(cells, threshold=75.0)

        exact = [m for m in matches if m.match_type == "Exact"]
        assert len(exact) >= 1  # "Bangalore" ↔ "Bangalore", "Japan" ↔ "Japan", etc.

    def test_finds_near_cell_matches(self):
        contents = _make_test_workbook()
        _, cells = parse_excel_file("test.xlsx", contents)
        matches = compare_cells(cells, threshold=75.0)

        near = [m for m in matches if m.match_type == "Near"]
        # "Bangalore" ↔ "Mangalore" should be a near match
        bangalore_mangalore = [
            m for m in near
            if ("Bangalore" in m.original_text and "Mangalore" in m.duplicate_text)
            or ("Mangalore" in m.original_text and "Bangalore" in m.duplicate_text)
        ]
        assert len(bangalore_mangalore) >= 1


class TestRunCrossComparison:
    def test_returns_both_levels(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)], threshold=75.0
        )
        assert len(row_matches) > 0
        assert len(cell_matches) > 0

    def test_can_disable_row_compare(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)],
            threshold=75.0,
            do_row_compare=False,
        )
        assert len(row_matches) == 0
        assert len(cell_matches) > 0

    def test_can_disable_cell_compare(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)],
            threshold=75.0,
            do_cell_compare=False,
        )
        assert len(row_matches) > 0
        assert len(cell_matches) == 0

    def test_multi_file_comparison(self):
        """Two separate files should be compared against each other."""
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.append(["Query"])
        ws1.append(["Japan 2025 trip photos"])
        buf1 = io.BytesIO()
        wb1.save(buf1)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["Query"])
        ws2.append(["Japan 2025 trip photos"])
        buf2 = io.BytesIO()
        wb2.save(buf2)

        row_matches, cell_matches = run_cross_comparison([
            ("file1.xlsx", buf1.getvalue()),
            ("file2.xlsx", buf2.getvalue()),
        ])

        exact_rows = [m for m in row_matches if m.match_type == "Exact"]
        assert len(exact_rows) >= 1


class TestReportGeneration:
    def test_generates_report_bytes(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)], threshold=75.0
        )
        report = generate_comparison_report(row_matches, cell_matches)
        assert isinstance(report, bytes)
        assert len(report) > 0

    def test_report_has_correct_sheets(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)], threshold=75.0
        )
        report = generate_comparison_report(row_matches, cell_matches)

        wb = load_workbook(io.BytesIO(report))
        assert "Row Duplicates" in wb.sheetnames
        assert "Cell Duplicates" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        wb.close()


class TestColoredWorkbook:
    def test_generates_colored_bytes(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)], threshold=75.0
        )
        colored = generate_colored_workbook(contents, row_matches, cell_matches)
        assert isinstance(colored, bytes)
        assert len(colored) > 0

    def test_colored_workbook_has_legend(self):
        contents = _make_test_workbook()
        row_matches, cell_matches = run_cross_comparison(
            [("test.xlsx", contents)], threshold=75.0
        )
        colored = generate_colored_workbook(contents, row_matches, cell_matches)

        wb = load_workbook(io.BytesIO(colored))
        ws = wb[wb.sheetnames[0]]
        # Find legend by checking for "Color Index" text
        found_legend = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Color Index":
                    found_legend = True
                    break
        assert found_legend
        wb.close()


# Need this import for test_report_has_correct_sheets
from openpyxl import load_workbook


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
