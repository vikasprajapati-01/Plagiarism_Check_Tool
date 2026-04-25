"""
Cross-cell & cross-row comparison for detecting duplicates across
Excel workbooks. Supports exact + fuzzy (Levenshtein) matching.
"""

import io
from dataclasses import dataclass, field
from itertools import combinations
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.preprocessor import preprocess_text
from app.services.fuzzy_match import levenshtein_similarity


# ── Data models ───────────────────────────────────────────────────────────────

@dataclass
class CellRef:
    file_name: str
    sheet_name: str
    row: int
    col: int
    col_letter: str
    raw_value: str
    cleaned_value: str = ""

    def __post_init__(self):
        self.cleaned_value = preprocess_text(self.raw_value) if self.raw_value else ""

    @property
    def label(self):
        return f"{self.file_name}-{self.col_letter}{self.row}"


@dataclass
class RowRef:
    file_name: str
    sheet_name: str
    row: int
    cells: List[CellRef] = field(default_factory=list)
    combined_raw: str = ""
    combined_cleaned: str = ""

    def __post_init__(self):
        self.combined_raw = " | ".join(c.raw_value for c in self.cells if c.raw_value)
        self.combined_cleaned = preprocess_text(self.combined_raw)

    @property
    def label(self):
        return f"{self.file_name}-Row {self.row}"


@dataclass
class MatchPair:
    original_label: str
    duplicate_label: str
    original_text: str
    duplicate_text: str
    match_type: str       # "Exact" / "Near"
    similarity: float     # 0-100
    level: str            # "Row" / "Cell"
    original_sheet: str = ""
    original_row: int = 0
    original_col: int = 0
    duplicate_sheet: str = ""
    duplicate_row: int = 0
    duplicate_col: int = 0


# ── Constants ─────────────────────────────────────────────────────────────────

FILL_EXACT_ROW  = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
FILL_NEAR_ROW   = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
FILL_EXACT_CELL = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
FILL_NEAR_CELL  = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER   = Border(left=Side("thin"), right=Side("thin"),
                   top=Side("thin"),  bottom=Side("thin"))

_SERIAL_HEADERS = {
    "s. no.", "s.no.", "s. no", "sno", "s no", "#",
    "sr", "sr.", "sr. no.", "sl. no.", "sl.no.", "sl no", "no.", "no",
}

_DEFAULT_TARGET = "query"


# ── Column discovery ──────────────────────────────────────────────────────────

def get_available_columns(files):
    """Return {\"file > sheet\": [col_headers]} for all uploaded files."""
    result = {}
    for fname, contents in files:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            hdrs = [str(v).strip() for v in next(ws.iter_rows(max_row=1, values_only=True)) if v]
            if hdrs:
                result[f"{fname} > {sname}"] = hdrs
        wb.close()
    return result


def _find_col_index(headers, col_name):
    """Find 1-indexed column position for col_name (case-insensitive)."""
    if not col_name:
        return None
    target = col_name.strip().lower()
    for i, h in enumerate(headers):
        if h == target:
            return i + 1
    return None


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_excel_file(file_name, contents, target_column=None):
    """Parse all sheets. Returns (rows, cells).

    target_column: "auto" = detect Query col, specific name = that col only, None = all.
    """
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    all_rows, all_cells = [], []

    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = []
        skip_cols = set()
        target_idx = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            if row_idx == 1:
                for ci, cell in enumerate(row, 1):
                    hval = str(cell.value).strip().lower() if cell.value else ""
                    headers.append(hval)
                    if hval in _SERIAL_HEADERS:
                        skip_cols.add(ci)

                # resolve which column to target
                if target_column:
                    if target_column.lower() == "auto":
                        target_idx = _find_col_index(headers, _DEFAULT_TARGET)
                    else:
                        target_idx = _find_col_index(headers, target_column)
                continue

            row_cells = []
            for ci, cell in enumerate(row, 1):
                if ci in skip_cols:
                    continue
                if target_idx is not None and ci != target_idx:
                    continue

                val = str(cell.value).strip() if cell.value is not None else ""
                if not val or val.lower() == "none":
                    continue

                ref = CellRef(file_name, sname, row_idx, ci,
                              get_column_letter(ci), val)
                row_cells.append(ref)
                all_cells.append(ref)

            if row_cells:
                all_rows.append(RowRef(file_name, sname, row_idx, row_cells))

    wb.close()
    return all_rows, all_cells


# ── Comparison engine ─────────────────────────────────────────────────────────

def _make_match(a_label, b_label, a_text, b_text, mtype, sim, level,
                a_sheet="", a_row=0, a_col=0, b_sheet="", b_row=0, b_col=0):
    return MatchPair(a_label, b_label, a_text, b_text, mtype, sim, level,
                     a_sheet, a_row, a_col, b_sheet, b_row, b_col)


def compare_rows(rows, threshold=75.0):
    """Pairwise row comparison (exact + fuzzy)."""
    matches = []
    for i, j in combinations(range(len(rows)), 2):
        a, b = rows[i], rows[j]
        if a.file_name == b.file_name and a.sheet_name == b.sheet_name and a.row == b.row:
            continue
        if not a.combined_cleaned or not b.combined_cleaned:
            continue

        if a.combined_cleaned == b.combined_cleaned:
            matches.append(_make_match(
                a.label, b.label, a.combined_raw, b.combined_raw,
                "Exact", 100.0, "Row", a.sheet_name, a.row, 0, b.sheet_name, b.row, 0))
        else:
            sim = levenshtein_similarity(a.combined_cleaned, b.combined_cleaned) * 100
            if sim >= threshold:
                matches.append(_make_match(
                    a.label, b.label, a.combined_raw, b.combined_raw,
                    "Near", round(sim, 1), "Row", a.sheet_name, a.row, 0, b.sheet_name, b.row, 0))
    return matches


def compare_cells(cells, threshold=75.0):
    """Pairwise cell comparison. Skips values shorter than 3 chars."""
    matches = []
    for i, j in combinations(range(len(cells)), 2):
        a, b = cells[i], cells[j]
        if (a.file_name == b.file_name and a.sheet_name == b.sheet_name
                and a.row == b.row and a.col == b.col):
            continue
        if len(a.cleaned_value) < 3 or len(b.cleaned_value) < 3:
            continue

        if a.cleaned_value == b.cleaned_value:
            matches.append(_make_match(
                a.label, b.label, a.raw_value, b.raw_value,
                "Exact", 100.0, "Cell", a.sheet_name, a.row, a.col, b.sheet_name, b.row, b.col))
        else:
            sim = levenshtein_similarity(a.cleaned_value, b.cleaned_value) * 100
            if sim >= threshold:
                matches.append(_make_match(
                    a.label, b.label, a.raw_value, b.raw_value,
                    "Near", round(sim, 1), "Cell", a.sheet_name, a.row, a.col, b.sheet_name, b.row, b.col))
    return matches


# ── Main entry point ──────────────────────────────────────────────────────────

def run_cross_comparison(files, threshold=75.0, do_row_compare=True,
                         do_cell_compare=True, target_column="auto"):
    """Parse files and run cross-comparison. Returns (row_matches, cell_matches)."""
    all_rows, all_cells = [], []
    for fname, contents in files:
        rows, cells = parse_excel_file(fname, contents, target_column)
        all_rows.extend(rows)
        all_cells.extend(cells)

    row_matches = compare_rows(all_rows, threshold) if do_row_compare else []
    cell_matches = compare_cells(all_cells, threshold) if do_cell_compare else []

    row_matches.sort(key=lambda m: m.similarity, reverse=True)
    cell_matches.sort(key=lambda m: m.similarity, reverse=True)
    return row_matches, cell_matches


# ── Report generation ─────────────────────────────────────────────────────────

def generate_comparison_report(row_matches, cell_matches):
    """Generate styled .xlsx report with Row Duplicates, Cell Duplicates, Summary sheets."""
    wb = Workbook()

    ws_rows = wb.active
    ws_rows.title = "Row Duplicates"
    _write_matches_sheet(ws_rows, row_matches)

    ws_cells = wb.create_sheet("Cell Duplicates")
    _write_matches_sheet(ws_cells, cell_matches)

    ws_sum = wb.create_sheet("Summary")
    _write_summary(ws_sum, row_matches, cell_matches)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_matches_sheet(ws, matches):
    headers = ["Original", "Duplicate", "Type", "Similarity (%)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.border = _HDR_FILL, _HDR_FONT, _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    exact_bg = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    near_bg  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for ri, m in enumerate(matches, 2):
        ws.cell(ri, 1, m.original_label).border = _BORDER
        ws.cell(ri, 2, m.duplicate_label).border = _BORDER

        tc = ws.cell(ri, 3, m.match_type)
        tc.border, tc.alignment = _BORDER, Alignment(horizontal="center")
        tc.fill = exact_bg if m.match_type == "Exact" else near_bg

        sc = ws.cell(ri, 4, m.similarity)
        sc.border, sc.number_format = _BORDER, "0"
        sc.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.freeze_panes = "A2"


def _write_summary(ws, row_m, cell_m):
    ws.cell(1, 1, "Cross-Comparison Summary").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:D1")

    lines = [
        ("", ""),
        ("Row-to-Row Duplicates", ""),
        ("  Total pairs found", len(row_m)),
        ("  Exact matches", sum(1 for m in row_m if m.match_type == "Exact")),
        ("  Near matches",  sum(1 for m in row_m if m.match_type == "Near")),
        ("", ""),
        ("Cell-to-Cell Duplicates", ""),
        ("  Total pairs found", len(cell_m)),
        ("  Exact matches", sum(1 for m in cell_m if m.match_type == "Exact")),
        ("  Near matches",  sum(1 for m in cell_m if m.match_type == "Near")),
        ("", ""),
        ("Overall", ""),
        ("  Total duplicate pairs", len(row_m) + len(cell_m)),
    ]
    for ri, (label, val) in enumerate(lines, 3):
        lc = ws.cell(ri, 1, label)
        if label and not label.startswith(" "):
            lc.font = Font(bold=True, size=11)
        if val != "":
            ws.cell(ri, 2, val)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


# ── Color-coded workbook ─────────────────────────────────────────────────────

def generate_colored_workbook(contents, row_matches, cell_matches):
    """Return a copy of the input workbook with duplicate cells highlighted."""
    wb = load_workbook(io.BytesIO(contents))

    # build lookup dicts
    row_colors = {}
    for m in row_matches:
        fill = FILL_EXACT_ROW if m.match_type == "Exact" else FILL_NEAR_ROW
        for key in [(m.original_sheet, m.original_row), (m.duplicate_sheet, m.duplicate_row)]:
            if key not in row_colors or m.match_type == "Exact":
                row_colors[key] = fill

    cell_colors = {}
    for m in cell_matches:
        fill = FILL_EXACT_CELL if m.match_type == "Exact" else FILL_NEAR_CELL
        for key in [(m.original_sheet, m.original_row, m.original_col),
                    (m.duplicate_sheet, m.duplicate_row, m.duplicate_col)]:
            if key not in cell_colors or m.match_type == "Exact":
                cell_colors[key] = fill

    # apply colors
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=2):
            ridx = row[0].row
            if (sname, ridx) in row_colors:
                for cell in row:
                    cell.fill = row_colors[(sname, ridx)]
            else:
                for cell in row:
                    key = (sname, ridx, cell.column)
                    if key in cell_colors:
                        cell.fill = cell_colors[key]

    # legend
    ws0 = wb[wb.sheetnames[0]]
    lcol = ws0.max_column + 2
    ws0.cell(1, lcol, "Color Index").font = Font(bold=True, size=11)
    for i, (txt, fill) in enumerate([
        ("Exact Row Duplicates",  FILL_EXACT_ROW),
        ("Near Row Duplicates",   FILL_NEAR_ROW),
        ("Exact Cell Duplicates", FILL_EXACT_CELL),
        ("Near Cell Duplicates",  FILL_NEAR_CELL),
    ], 2):
        c = ws0.cell(i, lcol, txt)
        c.fill, c.font = fill, Font(size=10)
    ws0.column_dimensions[get_column_letter(lcol)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
