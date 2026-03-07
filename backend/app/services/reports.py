"""
Plagiarism Detection Report Generator

Accepts detection results and produces styled, downloadable .xlsx reports
using pandas + openpyxl. Supports results from all three detection layers
(exact, fuzzy, semantic).

Usage:
    from app.services.reports import generate_report, generate_report_bytes

    results = [
        DetectionResult(
            text="Samsung Galaxy S23 Ultra",
            is_duplicate=True,
            similarity_scores={"levenshtein": 0.94, "cosine": 0.97},
            source="Samsung Galxy S23 Ultra",
            risk_level="high",
        ),
    ]

    # Save to file
    path = generate_report(results, output_path="report.xlsx")

    # Or get bytes (for FastAPI streaming response)
    file_bytes = generate_report_bytes(results)
"""

import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ==============================================================================
# DATA MODEL
# ==============================================================================


@dataclass
class DetectionResult:
    """Single plagiarism detection result entry.

    Attributes:
        text: The original text that was checked.
        is_duplicate: Whether the text was flagged as a duplicate.
        similarity_scores: Dict of algorithm → score, e.g.
            {"levenshtein": 0.94, "jaccard": 0.85, "cosine": 0.97}
        source: The matched text or source it was compared against.
        risk_level: "high", "medium", "low", or "none".
        detection_method: Which layer flagged it (exact/fuzzy/semantic).
        notes: Optional additional context.
    """

    text: str
    is_duplicate: bool
    similarity_scores: Dict[str, float] = field(default_factory=dict)
    source: Optional[str] = None
    risk_level: str = "none"
    detection_method: Optional[str] = None
    notes: Optional[str] = None


# ==============================================================================
# RISK LEVEL HELPERS
# ==============================================================================

_RISK_ORDER = {"high": 3, "medium": 2, "low": 1, "none": 0}


def classify_risk(score: float) -> str:
    """Classify a similarity score into a risk level.

    Args:
        score: Similarity score between 0.0 and 1.0.

    Returns:
        "high" if score >= 0.90, "medium" if >= 0.75, "low" if >= 0.50,
        otherwise "none".
    """
    if score >= 0.90:
        return "high"
    if score >= 0.75:
        return "medium"
    if score >= 0.50:
        return "low"
    return "none"


# ==============================================================================
# REPORT GENERATION
# ==============================================================================

# Style constants
_RISK_COLORS = {
    "high": "FF4C4C",     # Red
    "medium": "FFB347",   # Orange
    "low": "FFD700",      # Gold
    "none": "90EE90",     # Light green
}

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") if _OPENPYXL_AVAILABLE else None
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11) if _OPENPYXL_AVAILABLE else None
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
) if _OPENPYXL_AVAILABLE else None


def _build_dataframe(results: List[DetectionResult]) -> pd.DataFrame:
    """Convert detection results into a structured DataFrame."""
    rows = []
    for idx, r in enumerate(results, start=1):
        # Flatten similarity scores into a readable string
        scores_str = ", ".join(
            f"{algo}: {score:.3f}" for algo, score in r.similarity_scores.items()
        ) if r.similarity_scores else "—"

        # Pick the best (highest) score for the "Best Score" column
        best_score = max(r.similarity_scores.values()) if r.similarity_scores else 0.0

        rows.append(
            {
                "#": idx,
                "Text": r.text,
                "Duplicate?": "Yes" if r.is_duplicate else "No",
                "Risk Level": r.risk_level.capitalize(),
                "Best Score": round(best_score, 4),
                "Similarity Scores": scores_str,
                "Matched Source": r.source or "—",
                "Detection Method": (r.detection_method or "—").capitalize(),
                "Notes": r.notes or "",
            }
        )

    return pd.DataFrame(rows)


def _style_worksheet(ws, df: pd.DataFrame) -> None:
    """Apply professional styling to the worksheet."""
    if not _OPENPYXL_AVAILABLE:
        return

    num_cols = len(df.columns) + 1  # +1 because openpyxl is 1-indexed

    # --- Title row ---
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Plagiarism Detection Report"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Subtitle with timestamp ---
    ws.insert_rows(2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    subtitle_cell.font = Font(italic=True, size=10, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center")

    # --- Empty spacer row ---
    ws.insert_rows(3)
    header_row = 4  # data headers now start at row 4

    # --- Style header row ---
    for col_idx in range(1, num_cols):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    # --- Style data rows ---
    data_start = header_row + 1
    data_end = data_start + len(df) - 1

    for row_idx in range(data_start, data_end + 1):
        for col_idx in range(1, num_cols):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=col_idx in (2, 6, 7, 9),  # wrap long text columns
            )

        # Color the "Risk Level" column (column 4 in 1-indexed)
        risk_cell = ws.cell(row=row_idx, column=4)
        risk_key = (risk_cell.value or "none").lower()
        color = _RISK_COLORS.get(risk_key, "FFFFFF")
        risk_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        risk_cell.font = Font(bold=True)

        # Color the "Duplicate?" column (column 3)
        dup_cell = ws.cell(row=row_idx, column=3)
        if dup_cell.value == "Yes":
            dup_cell.font = Font(bold=True, color="CC0000")
        else:
            dup_cell.font = Font(color="228B22")

    # --- Auto-fit column widths ---
    column_widths = {
        1: 5,    # #
        2: 45,   # Text
        3: 12,   # Duplicate?
        4: 13,   # Risk Level
        5: 12,   # Best Score
        6: 40,   # Similarity Scores
        7: 40,   # Matched Source
        8: 18,   # Detection Method
        9: 30,   # Notes
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze panes below header ---
    ws.freeze_panes = ws.cell(row=data_start, column=1)

    # --- Summary section at bottom ---
    summary_row = data_end + 2
    total = len(df)
    duplicates = sum(1 for _, r in df.iterrows() if r["Duplicate?"] == "Yes")
    high_risk = sum(1 for _, r in df.iterrows() if r["Risk Level"] == "High")

    summary_items = [
        ("Total Texts Analyzed:", total),
        ("Duplicates Found:", duplicates),
        ("High Risk Items:", high_risk),
        ("Plagiarism Rate:", f"{(duplicates / total * 100):.1f}%" if total > 0 else "0%"),
    ]

    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(
        start_row=summary_row, start_column=1,
        end_row=summary_row, end_column=2,
    )

    for i, (label, value) in enumerate(summary_items):
        row = summary_row + 1 + i
        label_cell = ws.cell(row=row, column=1)
        value_cell = ws.cell(row=row, column=2)
        label_cell.value = label
        label_cell.font = Font(bold=True)
        value_cell.value = value


def generate_report(
    results: List[DetectionResult],
    output_path: str = "plagiarism_report.xlsx",
    sheet_name: str = "Detection Results",
) -> str:
    """
    Generate a styled .xlsx plagiarism report and save to disk.

    Args:
        results: List of DetectionResult objects.
        output_path: File path for the output .xlsx file.
        sheet_name: Name of the Excel sheet.

    Returns:
        The absolute path to the generated report file.

    Example:
        path = generate_report(results, output_path="report.xlsx")
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    df = _build_dataframe(results)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        ws = writer.sheets[sheet_name]
        _style_worksheet(ws, df)

    return output_path


def generate_report_bytes(
    results: List[DetectionResult],
    sheet_name: str = "Detection Results",
) -> bytes:
    """
    Generate a styled .xlsx report and return it as raw bytes.

    Ideal for FastAPI streaming responses:

        from fastapi.responses import StreamingResponse

        report_bytes = generate_report_bytes(results)
        return StreamingResponse(
            io.BytesIO(report_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=plagiarism_report.xlsx"},
        )

    Args:
        results: List of DetectionResult objects.
        sheet_name: Name of the Excel sheet.

    Returns:
        Bytes of the .xlsx file.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    df = _build_dataframe(results)
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        ws = writer.sheets[sheet_name]
        _style_worksheet(ws, df)

    return buffer.getvalue()
