"""Combined report endpoint — POST /reports/combined.

Accepts a PipelineRunResult-style JSON body and returns a 3-sheet Excel file.
"""

import io
import json
import logging
import re
import zipfile
from typing import Dict, List, Optional, Set, Tuple

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter()

_EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

# ── Colour constants ──────────────────────────────────────────────────────────
_HEADER_COLOR = "1F4E79"
_ROW_EXACT_COLOR = "FFCCCC"
_ROW_NEAR_COLOR = "FFF2CC"
_ROW_PLAG_YES_COLOR = "FFCCCC"
_ROW_PLAG_NO_COLOR = "C6EFCE"
_AI_RED = "FFCCCC"
_AI_ORANGE = "FFE0CC"
_AI_YELLOW = "FFF2CC"


# ── Inline request schema (pipeline result shape) ─────────────────────────────

class CombinedReportRequest(BaseModel):
    pipeline_id: str
    status: str = "completed"
    comparison_scope: str = "both"
    summary: dict
    row_duplicates: List[dict]
    cell_duplicates: List[dict]
    web_ai_results: List[dict]
    color_report: bool = False


def _parse_row_key(label: str) -> Optional[Tuple[str, int]]:
    if not label:
        return None
        
    parts = label.split("-Row")
    if len(parts) >= 2:
        row_part = parts[1].split("-")[0]
        try:
            return (parts[0], int(row_part))
        except ValueError:
            pass

    m = re.search(r"-(?:Row\s+(\d+)|([A-Z]+)(\d+))$", label)
    if m:
        row_num = int(m.group(1) or m.group(3))
        prefix = label[:m.start()]
        return (prefix, row_num)
    return (label, 0)


def _unique_rows_from_pairs(pairs: list, type_filter: str = None) -> set:
    seen = set()
    for pair in pairs:
        if type_filter and pair.get("type") != type_filter:
            continue
        for label in [pair.get("original", ""), pair.get("duplicate", "")]:
            key = _parse_row_key(label)
            if key:
                seen.add(key)
    return seen


def generate_pipeline_report(
    pipeline_id: str,
    row_duplicates: list,
    cell_duplicates: list,
    web_ai_results: list,
    color_report: bool = False,
    summary: Optional[dict] = None,
) -> bytes:
    def _build_pair_df(pairs: List[dict]) -> "pd.DataFrame":
        rows = [
            {
                "Original": row.get("original") or "",
                "Duplicate": row.get("duplicate") or "",
                "Type": row.get("type") or "",
                "Similarity (%)": row.get("similarity_pct", 0.0),
            }
            for row in pairs
        ]
        return pd.DataFrame(rows, columns=["Original", "Duplicate", "Type", "Similarity (%)"])

    def _build_web_ai_df(items: List[dict]) -> "pd.DataFrame":
        rows = [
            {
                "Original": row.get("original") or "",
                "Plagiarised": row.get("plagiarised") or "No",
                "Source": row.get("source") or "N/A",
                "AI Detected (%)": row.get("ai_detected_pct", 0.0),
            }
            for row in items
        ]
        df = pd.DataFrame(
            rows,
            columns=["Original", "Plagiarised", "Source", "AI Detected (%)"],
        )
        if not df.empty:
            df["Source"] = df["Source"].replace("", "N/A").fillna("N/A")
        return df

    df_rows = _build_pair_df(row_duplicates)
    df_cells = _build_pair_df(cell_duplicates)
    df_web_ai = _build_web_ai_df(web_ai_results)

    column_widths = {
        "Original": 55,
        "Duplicate": 55,
        "Type": 10,
        "Similarity (%)": 16,
        "Plagiarised": 14,
        "Source": 45,
        "AI Detected (%)": 18,
    }

    sheets = [
        ("Row-to-Row", df_rows),
        ("Cell-to-Cell", df_cells),
        ("AI-Plagiarism", df_web_ai),
    ]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            ws = writer.sheets[sheet_name]

            header_row = 1
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            data_start = 2
            data_end = data_start + len(df) - 1
            if len(df) > 0:
                for row_idx in range(data_start, data_end + 1):
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

            for col_idx, col_name in enumerate(df.columns, start=1):
                width = column_widths.get(col_name)
                if width:
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

            if color_report and len(df) > 0:
                if sheet_name in {"Row-to-Row", "Cell-to-Cell"}:
                    type_idx = list(df.columns).index("Type") + 1
                    for row_idx in range(data_start, data_end + 1):
                        type_value = ws.cell(row=row_idx, column=type_idx).value
                        if type_value == "Exact":
                            fill = PatternFill(start_color=_ROW_EXACT_COLOR, end_color=_ROW_EXACT_COLOR, fill_type="solid")
                        elif type_value == "Near":
                            fill = PatternFill(start_color=_ROW_NEAR_COLOR, end_color=_ROW_NEAR_COLOR, fill_type="solid")
                        else:
                            fill = None
                        if fill:
                            for col_idx in range(1, len(df.columns) + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = fill

                if sheet_name == "AI-Plagiarism":
                    plag_idx = list(df.columns).index("Plagiarised") + 1
                    ai_idx = list(df.columns).index("AI Detected (%)") + 1
                    for row_idx in range(data_start, data_end + 1):
                        plag_value = ws.cell(row=row_idx, column=plag_idx).value
                        if plag_value == "Yes":
                            row_fill = PatternFill(start_color=_ROW_PLAG_YES_COLOR, end_color=_ROW_PLAG_YES_COLOR, fill_type="solid")
                        elif plag_value == "No":
                            row_fill = PatternFill(start_color=_ROW_PLAG_NO_COLOR, end_color=_ROW_PLAG_NO_COLOR, fill_type="solid")
                        else:
                            row_fill = None

                        if row_fill:
                            for col_idx in range(1, len(df.columns) + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = row_fill

                        ai_cell = ws.cell(row=row_idx, column=ai_idx)
                        try:
                            ai_value = float(ai_cell.value or 0)
                        except (TypeError, ValueError):
                            ai_value = 0

                        ai_fill = None
                        if ai_value >= 80:
                            ai_fill = PatternFill(start_color=_AI_RED, end_color=_AI_RED, fill_type="solid")
                        elif ai_value >= 50:
                            ai_fill = PatternFill(start_color=_AI_ORANGE, end_color=_AI_ORANGE, fill_type="solid")
                        elif ai_value >= 20:
                            ai_fill = PatternFill(start_color=_AI_YELLOW, end_color=_AI_YELLOW, fill_type="solid")

                        if ai_fill:
                            ai_cell.fill = ai_fill

            ws.freeze_panes = "A2"

        # ── Create Sheet 4: Risk Summary ─────────────────────────────────────────
        wb = writer.book
        ws = wb.create_sheet("Risk Summary")
        ws.sheet_view.showGridLines = True

        # Write Headers
        headers_s4 = [
            "Detection Method",
            "Flagged Pairs",
            "Unique Rows Flagged",
            "Total Rows",
            "Flag Rate",
            "Risk Level",
        ]
        for col_idx, h in enumerate(headers_s4, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=_HEADER_COLOR,
                end_color=_HEADER_COLOR,
                fill_type="solid",
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 25

        total_rows = summary.get("total_rows", 0) if summary else 0

        all_duplicate_pairs = list(row_duplicates) + list(cell_duplicates)

        exact_pairs = [p for p in all_duplicate_pairs if p.get("type") == "Exact"]
        near_pairs  = [p for p in all_duplicate_pairs if p.get("type") == "Near"]

        exact_unique_rows = _unique_rows_from_pairs(exact_pairs)
        near_unique_rows = _unique_rows_from_pairs(near_pairs)
        
        ai_unique_rows = {
            _parse_row_key(r.get("original"))
            for r in web_ai_results
            if r.get("ai_detected_pct", 0.0) > 50.0
        } - {None}
        
        web_unique_rows = {
            _parse_row_key(r.get("original"))
            for r in web_ai_results
            if r.get("plagiarised") == "Yes"
        } - {None}

        semantic_pairs = 0
        ai_pairs = sum(1 for r in web_ai_results if r.get("ai_detected_pct", 0.0) > 50.0)
        web_pairs = sum(1 for r in web_ai_results if r.get("plagiarised") == "Yes")
        license_pairs = 0

        exact_unique = len(exact_pairs)
        near_unique = len(near_pairs)
        semantic_unique = 0
        ai_unique = ai_pairs
        web_unique = web_pairs
        license_unique = 0

        def _get_risk_level(flagged_unique: int, total: int) -> Tuple[str, Optional[str]]:
            if total <= 0:
                return "Clean", "C6EFCE"
            pct = flagged_unique / total * 100
            if pct == 0:
                return "Clean", "C6EFCE"
            elif pct <= 33:
                return "Low", "C6EFCE"
            elif pct <= 66:
                return "Medium", "FFEB9C"
            else:
                return "High", "FFC7CE"

        exact_risk, exact_risk_color = _get_risk_level(exact_unique, total_rows)
        near_risk, near_risk_color = _get_risk_level(near_unique, total_rows)
        ai_risk, ai_risk_color = _get_risk_level(ai_unique, total_rows)
        web_risk, web_risk_color = _get_risk_level(web_unique, total_rows)

        def _format_rate(flagged_unique: int, total: int) -> str:
            if total <= 0:
                return "N/A"
            return f"{(flagged_unique / total * 100):.2f}%"

        # Prepare section 1 rows
        rows_data = [
            ("Exact Duplicate", len(exact_pairs), exact_unique, total_rows, _format_rate(exact_unique, total_rows), exact_risk, exact_risk_color),
            ("Near Duplicate", len(near_pairs), near_unique, total_rows, _format_rate(near_unique, total_rows), near_risk, near_risk_color),
            ("Semantic Similar", semantic_pairs, semantic_unique, total_rows, "0.00%" if total_rows > 0 else "N/A", "Included in Near Duplicate", None),
            ("AI Generated", ai_pairs, ai_unique, total_rows, _format_rate(ai_unique, total_rows), ai_risk, ai_risk_color),
            ("Web Plagiarised", web_pairs, web_unique, total_rows, _format_rate(web_unique, total_rows), web_risk, web_risk_color),
            ("License Violation", license_pairs, license_unique, total_rows, "0.00%" if total_rows > 0 else "N/A", "Not separately tracked", None),
        ]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_num, (method, pairs, unique, total, rate, risk, color) in enumerate(rows_data, 2):
            # Column 1 (Detection Method)
            c1 = ws.cell(row=row_num, column=1, value=method)
            c1.alignment = left_align_wrap
            # Columns 2-5
            ws.cell(row=row_num, column=2, value=pairs)
            ws.cell(row=row_num, column=3, value=unique)
            ws.cell(row=row_num, column=4, value=total)
            ws.cell(row=row_num, column=5, value=rate)
            # Column 6 (Risk Level)
            c6 = ws.cell(row=row_num, column=6, value=risk)
            if color:
                c6.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Apply borders and alignment to all row cells
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = thin_border
                if col_idx > 1:
                    cell.alignment = center_align

        # Section 2 - Overall dataset summary (rows 8-10)
        # Row 8 is empty (already empty by default)
        
        # Row 9: Overall Dataset Summary Header
        ws.merge_cells("A9:F9")
        for col_idx in range(1, 7):
            cell = ws.cell(row=9, column=col_idx)
            cell.fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
            cell.border = thin_border
        
        hdr_cell = ws.cell(row=9, column=1, value="Overall Dataset Summary")
        hdr_cell.font = Font(bold=True, size=11, color="FFFFFF")
        hdr_cell.alignment = center_align
        ws.row_dimensions[9].height = 22

        # Row 10: Overall data
        overall_pairs = len(exact_pairs) + len(near_pairs) + semantic_pairs + ai_pairs + web_pairs + license_pairs
        overall_unique = overall_pairs
        overall_rate = _format_rate(overall_unique, total_rows)
        overall_risk, overall_risk_color = _get_risk_level(overall_unique, total_rows)

        ws.cell(row=10, column=1, value="All Methods Combined")
        ws.cell(row=10, column=2, value=overall_pairs)
        ws.cell(row=10, column=3, value=overall_unique)
        ws.cell(row=10, column=4, value=total_rows)
        ws.cell(row=10, column=5, value=overall_rate)
        
        c10_6 = ws.cell(row=10, column=6, value=overall_risk)
        if overall_risk_color:
            c10_6.fill = PatternFill(start_color=overall_risk_color, end_color=overall_risk_color, fill_type="solid")

        for col_idx in range(1, 7):
            cell = ws.cell(row=10, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align
            if col_idx == 1:
                cell.alignment = left_align_wrap

        # Section 3 - PRS Score (rows 11-12)
        # Row 11 is empty
        
        # Row 12
        lbl_cell = ws.cell(row=12, column=1, value="Plagiarism Risk Score (PRS)")
        lbl_cell.font = Font(bold=True, size=11)
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        total_for_prs = total_rows if total_rows > 0 else 1
        exact_rate = exact_unique / total_for_prs
        near_rate = near_unique / total_for_prs
        semantic_rate = 0.0
        ai_rate = ai_unique / total_for_prs
        web_rate = web_unique / total_for_prs
        license_rate = 0.0

        raw_prs = (
            exact_rate * 100.0 * 0.40
            + near_rate * 100.0 * 0.20
            + semantic_rate * 100.0 * 0.15
            + ai_rate * 100.0 * 0.10
            + web_rate * 100.0 * 0.10
            + license_rate * 100.0 * 0.05
        )
        prs = round(min(100.0, raw_prs), 1)
        
        score_cell = ws.cell(row=12, column=2, value=f"{prs:.1f} / 100")
        score_cell.font = Font(bold=True, size=11)
        score_cell.alignment = center_align
        
        if prs < 20:
            prs_bg = "C6EFCE"
        elif prs < 40:
            prs_bg = "FFEB9C"
        elif prs < 60:
            prs_bg = "FFE0CC"
        else:
            prs_bg = "FFC7CE"
            
        score_cell.fill = PatternFill(start_color=prs_bg, end_color=prs_bg, fill_type="solid")
        score_cell.border = thin_border
        lbl_cell.border = thin_border

        # Section 4 - Dataset Quality Assessment (rows 13-14)
        # Row 13 is empty
        
        # Row 14
        qa_lbl = ws.cell(row=14, column=1, value="Dataset Quality Assessment:")
        qa_lbl.font = Font(bold=True, size=11)
        qa_lbl.alignment = Alignment(horizontal="left", vertical="center")
        qa_lbl.border = thin_border

        if prs < 20:
            qa_text = "LOW RISK — Dataset appears clean for training use."
            qa_color = "375623"
        elif prs < 40:
            qa_text = "MEDIUM RISK — Minor issues found. Review flagged entries before training."
            qa_color = "7D6608"
        elif prs < 60:
            qa_text = "HIGH RISK — Significant issues found. Cleaning recommended before training."
            qa_color = "974706"
        else:
            qa_text = "CRITICAL RISK — Dataset has serious quality issues. Do not use for training without cleaning."
            qa_color = "9C0006"

        qa_val = ws.cell(row=14, column=2, value=qa_text)
        qa_val.font = Font(bold=True, size=11, color=qa_color)
        qa_val.alignment = Alignment(horizontal="left", vertical="center")
        qa_val.border = thin_border
        
        ws.merge_cells("B14:F14")
        for col_idx in range(2, 7):
            ws.cell(row=14, column=col_idx).border = thin_border

        # Column widths for Sheet 4
        s4_widths = {
            "A": 28,
            "B": 16,
            "C": 22,
            "D": 14,
            "E": 12,
            "F": 14,
        }
        for col_let, width in s4_widths.items():
            ws.column_dimensions[col_let].width = width

        # Freeze panes A2
        ws.freeze_panes = "A2"

    return buf.getvalue()


# ── Helpers for cleaned-output endpoint ──────────────────────────────────────

def _parse_label(
    label: str,
    known_files: Set[str],
) -> Optional[Tuple[str, str, int]]:
    """Parse a pipeline label into (file_name, sheet_name, excel_row_number).

    Label formats produced by cross_compare.py and pipeline_runner.py:
      Row  : "filename-SheetName-Row 5"   → row = 5
      Cell : "filename-SheetName-B5"      → row = 5
      Web  : "filename-SheetName-B5"      → row = 5  (same as cell)
    """
    # Row format: suffix is "-Row <digits>"
    row_match = re.search(r"-Row (\d+)$", label)
    if row_match:
        row_num = int(row_match.group(1))
        prefix = label[: row_match.start()]  # "file_name-sheet_name"
    else:
        # Cell/Web format: suffix is "-<LETTERS><digits>"
        cell_match = re.search(r"-([A-Z]+)(\d+)$", label)
        if not cell_match:
            return None
        row_num = int(cell_match.group(2))
        prefix = label[: cell_match.start()]  # "file_name-sheet_name"

    # Match prefix against known uploaded filenames (longest first to handle
    # filenames that themselves contain hyphens).
    for fname in sorted(known_files, key=len, reverse=True):
        if prefix == fname:
            return (fname, "", row_num)
        if prefix.startswith(fname + "-"):
            sheet = prefix[len(fname) + 1 :]
            return (fname, sheet, row_num)

    return None  # could not resolve to a known file


# ── POST /cleaned ─────────────────────────────────────────────────────────────

@router.post("/cleaned")
async def cleaned_report(
    files: List[UploadFile] = File(default=[]),
    row_duplicates: str = Form(default="[]"),
    cell_duplicates: str = Form(default="[]"),
    web_ai_results: str = Form(default="[]"),
):
    """Return cleaned .xlsx file(s) with duplicate / plagiarised rows removed.

    Rules:
      - row_duplicates / cell_duplicates : the *duplicate* row is deleted;
        the *original* row is kept.
      - web_ai_results : the *original* row is deleted only when
        plagiarised == "Yes".  AI-only entries are left untouched.
      - Only .xlsx files are processed.  CSV / TXT are ignored.
      - Row 1 (header) is never deleted.
    """
    if not _EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="openpyxl / pandas not installed.",
        )

    # ── Parse JSON form fields ────────────────────────────────────────────────
    try:
        row_dups: List[dict] = json.loads(row_duplicates)
        cell_dups: List[dict] = json.loads(cell_duplicates)
        web_ai: List[dict] = json.loads(web_ai_results)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=422, detail=f"Invalid JSON in form field: {exc}")

    # ── Read only .xlsx uploads ───────────────────────────────────────────────
    xlsx_files: List[Tuple[str, bytes]] = []
    for upload in files:
        fname = upload.filename or ""
        if fname.lower().endswith(".xlsx"):
            xlsx_files.append((fname, await upload.read()))

    if not xlsx_files:
        raise HTTPException(
            status_code=400,
            detail="No .xlsx files provided. Cleaned output only works with Excel (.xlsx) files.",
        )

    known_files: Set[str] = {fname for fname, _ in xlsx_files}

    # ── Build deletion map: { filename → { sheet → set of row numbers } } ────
    rows_to_delete: Dict[str, Dict[str, Set[int]]] = {}

    def _mark(label: str) -> None:
        parsed = _parse_label(label, known_files)
        if parsed is None:
            return
        fname, sheet, row_num = parsed
        rows_to_delete.setdefault(fname, {}).setdefault(sheet, set()).add(row_num)

    # Duplicate pairs → delete the *duplicate* side
    for pair in row_dups:
        _mark(pair.get("duplicate", ""))
    for pair in cell_dups:
        _mark(pair.get("duplicate", ""))

    # Web/AI → delete the *original* only when explicitly plagiarised
    for entry in web_ai:
        if str(entry.get("plagiarised", "")).strip().lower() == "yes":
            _mark(entry.get("original", ""))

    logger.info(
        "Cleaned report: %d xlsx file(s), deletion map=%s",
        len(xlsx_files),
        {k: {s: sorted(v) for s, v in d.items()} for k, d in rows_to_delete.items()},
    )

    # ── Process each file ─────────────────────────────────────────────────────
    cleaned_files: List[Tuple[str, bytes]] = []

    for fname, contents in xlsx_files:
        file_deletions = rows_to_delete.get(fname, {})

        try:
            wb = load_workbook(io.BytesIO(contents))
        except Exception as exc:
            logger.warning("Failed to open workbook %s: %s", fname, exc)
            continue

        for sname in wb.sheetnames:
            sheet_rows: Set[int] = file_deletions.get(sname, set())
            if not sheet_rows:
                continue
            ws = wb[sname]
            max_row = ws.max_row or 1
            # Delete in reverse order so earlier row indices stay valid
            for row_num in sorted(sheet_rows, reverse=True):
                if row_num > 1 and row_num <= max_row:  # never delete header (row 1)
                    ws.delete_rows(row_num)

        buf = io.BytesIO()
        wb.save(buf)
        base = fname.rsplit(".", 1)[0]
        cleaned_files.append((f"{base}_cleaned.xlsx", buf.getvalue()))

    if not cleaned_files:
        raise HTTPException(
            status_code=500,
            detail="Could not process any uploaded .xlsx file.",
        )

    # ── Return single file or zip ─────────────────────────────────────────────
    if len(cleaned_files) == 1:
        out_name, out_bytes = cleaned_files[0]
        return StreamingResponse(
            io.BytesIO(out_bytes),
            media_type=_EXCEL_MEDIA_TYPE,
            headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
        )

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for out_name, out_bytes in cleaned_files:
            zf.writestr(out_name, out_bytes)
    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="cleaned_files.zip"'},
    )


@router.post("/combined")
async def combined_report(body: CombinedReportRequest):
    """Generate a 3-sheet Excel report from a pipeline run result."""
    if not _EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="openpyxl or pandas not installed. Run: pip install openpyxl pandas",
        )
    report_bytes = generate_pipeline_report(
        pipeline_id=body.pipeline_id,
        row_duplicates=body.row_duplicates,
        cell_duplicates=body.cell_duplicates,
        web_ai_results=body.web_ai_results,
        color_report=body.color_report,
        summary=body.summary,
    )

    filename = f"pipeline_{body.pipeline_id[:8]}_report.xlsx"
    return StreamingResponse(
        io.BytesIO(report_bytes),
        media_type=_EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
