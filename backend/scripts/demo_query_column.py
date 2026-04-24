"""
Generate the 3-dataset sample + run cross-comparison targeting only the Query column.
Saves: sample_input.xlsx, query_report.xlsx, query_colored.xlsx
"""
import io, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from app.services.cross_compare import (
    get_available_columns,
    run_cross_comparison,
    generate_comparison_report,
    generate_colored_workbook,
)

out_dir = os.path.dirname(os.path.abspath(__file__))

# ── Build the 3-dataset workbook (from mentor's image) ────────────────────────
wb = Workbook()

ws1 = wb.active
ws1.title = "Dataset 1"
ws1.append(["S. No.", "Query", "Location", "Time"])
ws1.append([1,  "John playing foosball in Bangalore yesterday",                       "Bangalore",   "yesterday"])
ws1.append([2,  "The concert I watched last week in Delhi",                           "Delhi",       "last week"])
ws1.append([3,  "The performance I watched last week in Delhi",                       "Delhi",       "last week"])
ws1.append([4,  "Photos of 2019 BTS world tour in Los Angeles",                      "Los Angeles", "2019"])
ws1.append([5,  "Photos taken at the beach in Busan last summer",                    "Busan",       "last summer"])
ws1.append([6,  "Photos of cats taken on the streets of Ooty this month",            "Ooty",        "this month"])
ws1.append([7,  "Photos from when I went to the Holi party in Hyderabad in February","Hyderabad",   "February"])
ws1.append([8,  "Hyderabad Holi party photos from February",                         "Hyderabad",   "February"])
ws1.append([9,  "John playing foosball in Bangalore yesterday",                      "Bangalore",   "yesterday"])
ws1.append([10, "Japan 2025 trip photos",                                            "Japan",       "2025"])

ws2 = wb.create_sheet("Dataset 2")
ws2.append(["S. No.", "Query", "Location", "Time"])
ws2.append([1,  "Joel playing table tennis in Mangalore yesterday",                  "Mangalore",   "yesterday"])
ws2.append([2,  "The concert I watched last week in Chennai",                        "Chennai",     "last week"])
ws2.append([3,  "The friend I met in Mysuru last year",                              "Mysuru",      "last year"])
ws2.append([4,  "Tokyo cherry blossom picnic shots from late March",                 "Tokyo",       "late March"])
ws2.append([5,  "Group selfies from the Coachella music festival in Indio taken last week","Indio", "last week"])
ws2.append([6,  "Traditional tea ceremony photos taken in Kyoto last autumn",       "Kyoto",       "last autumn"])
ws2.append([7,  "Family dinner photos taken at a rooftop bistro in Bangkok last night","Bangkok",   "last night"])
ws2.append([8,  "Group photos from the carnival parade in Rio de Janeiro last week", "Hyderabad",   "February"])
ws2.append([9,  "Street food tour captures from Old Delhi in 2024",                  "Old Delhi",   "2024"])
ws2.append([10, "Japan 2025 trip photos",                                             "Japan",       "2025"])

ws3 = wb.create_sheet("Dataset 3")
ws3.append(["S. No.", "Query", "Location", "Time"])
ws3.append([1,  "2026 Jaipur trip photos",                                           "Jaipur",      "2026"])
ws3.append([2,  "Admit card for last month's JLPT exam in Mumbai",                   "Mumbai",      "last month"])
ws3.append([3,  "The photo exhibition I visited two weeks ago in Pune",              "Pune",        "two weeks ago"])
ws3.append([4,  "Last weekend's pilates workshop in Surat",                          "Surat",       "Last weekend"])
ws3.append([5,  "12 am on New Year's Day in Fukuoka",                                "Fukuoka",     "12.00 AM"])
ws3.append([6,  "Traditional tea ceremony photos taken in Nagoya last autumn",      "Nagoya",      "last autumn"])
ws3.append([7,  "Riya dancing at a wedding in Udaipur last summer",                 "Udaipur",     "last summer"])
ws3.append([8,  "Rohan visiting the Golden Temple in Amritsar last March",          "Amritsar",    "last March"])
ws3.append([9,  "Shimla snowfall videos from last December",                         "Shimla",      "last December"])
ws3.append([10, "Japan 2025 trip photos",                                             "Japan",       "2025"])

buf = io.BytesIO()
wb.save(buf)
input_bytes = buf.getvalue()

# Save input
input_path = os.path.join(out_dir, "sample_input.xlsx")
with open(input_path, "wb") as f:
    f.write(input_bytes)

# ── Step 1: Show available columns ────────────────────────────────────────────
print("=" * 70)
print("COLUMN DISCOVERY")
print("=" * 70)
columns = get_available_columns([("sample_input.xlsx", input_bytes)])
for sheet, headers in columns.items():
    print(f"  {sheet}: {headers}")

has_query = any(
    any(h.lower() == "query" for h in headers)
    for headers in columns.values()
)
print(f"\n  ✅ 'Query' column detected: {has_query}")

# ── Step 2: Run with auto-detect (only Query column) ─────────────────────────
print("\n" + "=" * 70)
print("COMPARISON — Query column only (auto-detected)")
print("=" * 70)

row_matches, cell_matches = run_cross_comparison(
    [("sample_input.xlsx", input_bytes)],
    threshold=75.0,
    target_column="auto",   # ← auto-detects "Query"
)

print(f"\nRow-to-Row matches: {len(row_matches)}")
for m in row_matches:
    tag = "✅ Exact" if m.match_type == "Exact" else "🟡 Near "
    print(f"  {tag}  {m.original_label}  ↔  {m.duplicate_label}  ({m.similarity}%)")

print(f"\nCell-to-Cell matches: {len(cell_matches)}")
for m in cell_matches:
    tag = "✅ Exact" if m.match_type == "Exact" else "🟡 Near "
    print(f"  {tag}  {m.original_label}  ↔  {m.duplicate_label}  ({m.similarity}%)")
    print(f"         '{m.original_text}'  ↔  '{m.duplicate_text}'")

# ── Step 3: Generate report + colored workbook ────────────────────────────────
report_bytes = generate_comparison_report(row_matches, cell_matches)
report_path = os.path.join(out_dir, "query_report.xlsx")
with open(report_path, "wb") as f:
    f.write(report_bytes)

colored_bytes = generate_colored_workbook(input_bytes, row_matches, cell_matches)
colored_path = os.path.join(out_dir, "query_colored.xlsx")
with open(colored_path, "wb") as f:
    f.write(colored_bytes)

print(f"\n{'=' * 70}")
print(f"📄 Input:    {input_path}")
print(f"📊 Report:   {report_path}  ({len(report_bytes):,} bytes)")
print(f"🎨 Colored:  {colored_path}  ({len(colored_bytes):,} bytes)")
print(f"{'=' * 70}")

# Open them
os.startfile(input_path)
os.startfile(report_path)
os.startfile(colored_path)
